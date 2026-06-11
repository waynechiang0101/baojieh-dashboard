#!/usr/bin/env python3
"""
康是美供應商平台 — 實銷資料自動抓取
每天自動登入、抓所有 P&G 廠編、合併成月報 xlsx

用法：python3 scripts/cosmed_fetch.py
輸出：~/Downloads/康是美實銷_YYYYMM.xlsx
"""
import os, json, re, base64, glob
from datetime import datetime
from playwright.sync_api import sync_playwright
import anthropic
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── 設定 ─────────────────────────────────────────
COSMED_USER = '860417111'
COSMED_PASS = 'Pj86041711'
ANTHROPIC_KEY = os.environ.get('ANTHROPIC_API_KEY', '')

# 價格表來源（條碼→零售價/成本）
PRICE_XLSX_PATTERN = os.path.expanduser('~/Downloads/康是美門市實銷資料*.xlsx')

def load_price_map():
    """從康是美實銷 xlsx 讀取 SKU 零售價和成本，以條碼為 key。"""
    files = sorted(glob.glob(PRICE_XLSX_PATTERN), key=os.path.getmtime, reverse=True)
    if not files:
        print('  ⚠ 找不到康是美實銷價格表')
        return {}
    path = files[0]
    print(f'  讀取價格表: {os.path.basename(path)}')
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sh = wb['2026實銷資料']
        price_map = {}
        hdr_row = None
        for i in range(1, 10):
            row = [str(sh.cell(i, j).value or '') for j in range(1, 12)]
            if '條碼' in row:
                hdr_row = i
                break
        if not hdr_row:
            return {}
        for i in range(hdr_row + 1, sh.max_row + 1):
            barcode = str(sh.cell(i, 5).value or '').strip()
            retail  = sh.cell(i, 10).value
            cost    = sh.cell(i, 9).value
            if barcode and (retail or cost):
                try:
                    price_map[barcode] = {
                        'retail': float(retail) if retail else 0.0,
                        'cost':   float(cost) if cost else 0.0
                    }
                except:
                    pass
        print(f'  ✓ 價格表: {len(price_map)} 筆 SKU')
        return price_map
    except Exception as e:
        print(f'  ⚠ 價格表讀取失敗: {e}')
        return {}

# 保留的 P&G 廠編（排除 085804/065189/203135）
PG_SUPPLIERS = [
    '8604171183',  # OLAY PRO X
    '8604171186',  # ORALB / BRAUN
    '8604171187',  # ARIEL
    '8604171189',  # GLT 吉列
    '8604171191',  # CREST
    '8604171192',  # OLAY 歐蕾
    '8604171193',  # WHSP 好自在
    '8604171199',  # HR 髮的食譜
]

BRAND_MAP = {
    '8604171183': 'OLAY',
    '8604171186': 'ORALB',
    '8604171187': 'ARIEL',
    '8604171189': 'GLT',
    '8604171191': 'CREST',
    '8604171192': 'OLAY',
    '8604171193': 'WHSP',
    '8604171199': 'HAIR',  # HR/HS/PNTN/PERT 全髮類
}

# 從品名判斷品牌
BRAND_KEYWORDS = [
    ('PAMPS',  ['幫寶適', 'PAMPERS']),
    ('WHSP',   ['whisper', '好自在', 'Whisper']),
    ('OLAY',   ['OLAY', '歐蕾', 'PRO X']),
    ('HS',     ['海倫仙度絲', 'Head & Shoulders']),
    ('PNTN',   ['潘婷', 'PANTENE']),
    ('PERT',   ['飛柔', 'PERT']),
    ('HR',     ['髮の食譜', '髮的食譜']),
    ('ARIEL',  ['ARIEL', 'Ariel']),
    ('LENOR',  ['蘭諾', 'LENOR', 'Lenor']),
    ('ORALB',  ['Oral-B', 'ORAL-B', '歐樂']),
    ('BRAUN',  ['BRAUN', '百靈']),
    ('CREST',  ['Crest', 'CREST']),
    ('GLT',    ['吉列', 'Gillette', 'GILLETTE']),
    ('FBRZ',   ['風倍清', 'Febreze']),
    ('BOLD',   ['BOLD']),
]

def detect_brand(name, default='OTHER'):
    for brand, keywords in BRAND_KEYWORDS:
        if any(kw in name for kw in keywords):
            return brand
    return default

# ── 驗證碼辨識 ────────────────────────────────────
def read_captcha(page):
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    captcha_bytes = page.locator('img[src*="securityImage"]').first.screenshot()
    b64 = base64.standard_b64encode(captcha_bytes).decode()
    resp = client.messages.create(
        model='claude-opus-4-5', max_tokens=20,
        messages=[{'role':'user','content':[
            {'type':'image','source':{'type':'base64','media_type':'image/png','data':b64}},
            {'type':'text','text':'這張圖片裡的驗證碼文字是什麼？注意英文大小寫，只回答驗證碼本身。'}
        ]}]
    )
    return resp.content[0].text.strip()

# ── 登入 ──────────────────────────────────────────
def login(page):
    for attempt in range(4):
        page.goto('https://scm.cosmed.com.tw/APCSM/loginOut', timeout=30000)
        captcha = read_captcha(page)
        print(f'  驗證碼: {captcha}')
        page.fill('input[name="loginUserId"]', COSMED_USER)
        page.fill('input[name="loginUserPp"]', COSMED_PASS)
        page.fill('input[name="securityId"]', captcha)
        page.click('button:has-text("登入")')
        page.wait_for_load_state('networkidle', timeout=15000)
        if 'home' in page.url:
            print('  登入成功!')
            return True
        print(f'  嘗試{attempt+1}失敗，重試...')
    return False

# ── 抓取單一廠編資料（只用頁碼翻頁，不動 page size）──────────────────────
def fetch_supplier(page, supplier_id):
    page.locator('select').first.select_option(supplier_id)
    page.locator('button:has-text("查詢")').click()
    page.wait_for_load_state('networkidle', timeout=20000)
    page.wait_for_timeout(800)

    # 抓表頭
    raw_headers = [th.inner_text().strip() for th in page.locator('table thead tr th').all()]
    seen = []; headers = []
    for h in raw_headers:
        if h not in seen:
            seen.append(h); headers.append(h)
    week_cols = [h for h in headers if '~' in h]

    all_rows = []
    seen_keys = set()

    def collect_rows():
        for tr in page.locator('table tbody tr').all():
            cells = [td.inner_text().strip() for td in tr.locator('td').all()]
            if len(cells) >= 6:
                key = f"{cells[3]}_{cells[4]}"
                if key not in seen_keys:
                    seen_keys.add(key)
                    all_rows.append(cells)

    collect_rows()

    # 取頁碼選項（第三個 select：1,2,3...）
    # 注意：不動第二個 select（page size），只翻頁碼
    while True:
        try:
            selects = page.locator('select').all()
            if len(selects) < 3:
                break
            page_sel = selects[2]
            opts = page_sel.locator('option').all()
            # 找目前選中的頁碼
            current = page_sel.evaluate('el => el.value')
            all_vals = [o.get_attribute('value') for o in opts]
            if not all_vals or current == all_vals[-1]:
                break
            # 選下一頁
            idx = all_vals.index(current)
            next_pg = all_vals[idx + 1]
            page_sel.select_option(next_pg)
            page.wait_for_timeout(2000)  # 背景請求太多，networkidle會timeout
            collect_rows()
        except:
            break

    return week_cols, all_rows

# ── 庫存數量查詢（康是美 DC：捷盟北倉+南倉+西園倉）─────────
def fetch_inventory_supplier(page, supplier_id):
    """單一廠編的 DC 庫存。欄位：No.|品號|條碼|品名|北倉休配|南倉休配|西園倉休配|
    訂單在途量|物流庫存|捷盟北倉|捷盟南倉|捷盟西園倉|庫存日期|供應商料號"""
    page.locator('select').first.select_option(supplier_id)
    page.locator('button:has-text("查詢")').click()
    page.wait_for_timeout(2000)

    all_rows, seen = [], set()
    def collect():
        for tr in page.locator('table tbody tr').all():
            c = [td.inner_text().strip() for td in tr.locator('td').all()]
            if len(c) >= 13 and c[1] not in seen:
                seen.add(c[1])
                all_rows.append(c)
    collect()
    while True:
        try:
            selects = page.locator('select').all()
            if len(selects) < 3: break
            page_sel = selects[2]
            vals = [o.get_attribute('value') for o in page_sel.locator('option').all()]
            cur = page_sel.evaluate('el => el.value')
            if not vals or cur == vals[-1]: break
            page_sel.select_option(vals[vals.index(cur)+1])
            page.wait_for_timeout(2000)
            collect()
        except: break
    return all_rows

def fetch_km_inventory(page):
    """全部 P&G 廠編的 DC 庫存，回傳 {條碼: {...}}（休配/在途/物流庫存/三倉）"""
    page.locator('a:has-text("銷售庫存")').click()
    page.wait_for_timeout(500)
    page.locator('a:has-text("庫存數量查詢")').click()
    page.wait_for_timeout(2500)

    inv = {}
    def n(v):
        try: return int(str(v).replace(',',''))
        except: return 0
    for sid in PG_SUPPLIERS:
        rows = fetch_inventory_supplier(page, sid)
        for c in rows:
            bc = c[2].split('\n')[0].strip()
            if not bc: continue
            inv[bc] = {
                'status': c[4][:1],          # 休配狀態碼（4=停售下架）
                'transit': n(c[7]),          # 訂單在途量
                'dc': n(c[8]),               # 物流庫存（三倉合計）
                'wh': [n(c[9]), n(c[10]), n(c[11])],  # 北/南/西園
            }
        print(f'    庫存 {sid}: {len(rows)}筆')
    print(f'  ✓ DC庫存 {len(inv)} SKU')
    return inv


# ── 主流程 ────────────────────────────────────────
def fetch_all_km_sell():
    """給 derp_fetch.py 呼叫，回傳康是美實銷彙總資料。
    回傳格式：
    {
      'weeks': ['20260601~20260607', ...],   # 週期欄位
      'by_brand': {
        'ARIEL': {'週期': 總件數, ...},
        'OLAY':  {...},
        ...
      },
      'by_sku': [
        {'brand':'ARIEL','name':'...','barcode':'...','weeks':[件數,...]},
        ...
      ]
    }
    """
    from collections import defaultdict

    # 讀價格表
    price_map = load_price_map()

    all_rows = []
    week_cols = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        print('  登入康是美...')
        if not login(page):
            print('  ❌ 登入失敗')
            browser.close()
            return None

        page.locator('a:has-text("銷售庫存")').click()
        page.wait_for_timeout(500)
        page.locator('a:has-text("銷售數量查詢")').click()
        page.wait_for_load_state('networkidle', timeout=15000)

        for sid in PG_SUPPLIERS:
            brand_default = BRAND_MAP.get(sid, 'OTHER')
            wc, rows = fetch_supplier(page, sid)
            if not week_cols and wc:
                week_cols = wc
            for row in rows:
                if len(row) >= 6:
                    name = row[5]
                    barcode = row[3] if len(row) > 3 else ''
                    brand = detect_brand(name, brand_default)
                    # 條碼可能多行（換行分隔），取第一個
                    bc_key = barcode.split('\n')[0].strip()
                    price_info = price_map.get(bc_key, {})
                    retail = price_info.get('retail', 0.0)
                    cost   = price_info.get('cost', 0.0)

                    week_vals = []
                    for w in row[6:6+len(week_cols)]:
                        try:
                            week_vals.append(int(w.replace(',', '')) if w else 0)
                        except:
                            week_vals.append(0)
                    all_rows.append({
                        'supplier_id': row[1] if len(row) > 1 else sid,
                        'brand':    brand,
                        'barcode':  barcode,
                        'item_no':  row[4] if len(row) > 4 else '',
                        'name':     name,
                        'retail':   retail,   # 零售價（0 = 無資料）
                        'cost':     cost,     # 成本含稅
                        'weeks':    week_vals,
                    })
            print(f'    {sid}: {len(rows)}筆')

        # DC 庫存（同一個 session 順手抓）
        dc_inv = {}
        try:
            dc_inv = fetch_km_inventory(page)
        except Exception as e:
            print(f'  ⚠ DC庫存抓取失敗（實銷不受影響）: {e}')

        browser.close()

    if not week_cols:
        return None

    # 彙總 by brand：件數 + 預估銷售額（件數 × 康是美進貨價 = 寶捷出貨價）
    by_brand_qty = defaultdict(lambda: [0] * len(week_cols))
    by_brand_amt = defaultdict(lambda: [0.0] * len(week_cols))
    no_price_count = 0

    for row in all_rows:
        for i, qty in enumerate(row['weeks']):
            if i < len(week_cols):
                by_brand_qty[row['brand']][i] += qty
                if row['cost'] > 0:
                    by_brand_amt[row['brand']][i] += qty * row['cost']
                else:
                    no_price_count += 1 if i == 0 else 0

    if no_price_count:
        print(f'  ⚠ {no_price_count} 筆 SKU 無進貨價，銷售額估算不含這些')

    print(f'  ✓ 康是美實銷: {len(all_rows)}筆 SKU, {len(week_cols)}週')
    return {
        'weeks':        week_cols,
        'by_brand_qty': {b: list(v) for b, v in by_brand_qty.items()},
        'by_brand_amt': {b: [round(v) for v in vals] for b, vals in by_brand_amt.items()},
        'by_sku': [{
            'brand':   s['brand'],
            'name':    s['name'][:30],
            'barcode': s['barcode'],
            'retail':  s['retail'],
            'cost':    s['cost'],
            'weeks':   s['weeks'],
            'amt_weeks': [round(q * s['cost']) if s['cost'] > 0 else 0
                          for q in s['weeks']],
            **({'dc': dc_inv[s['barcode'].split('\n')[0].strip()]['dc'],
                'transit': dc_inv[s['barcode'].split('\n')[0].strip()]['transit'],
                'kst': dc_inv[s['barcode'].split('\n')[0].strip()]['status']}
               if s['barcode'].split('\n')[0].strip() in dc_inv else {}),
        } for s in all_rows],
    }


def main():
    today = datetime.now()
    out_path = os.path.expanduser(f'~/Downloads/康是美實銷_{today.strftime("%Y%m")}.xlsx')

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        print('[登入康是美]')
        if not login(page):
            print('❌ 登入失敗')
            browser.close()
            return

        # 去銷售查詢
        page.locator('a:has-text("銷售庫存")').click()
        page.wait_for_timeout(500)
        page.locator('a:has-text("銷售數量查詢")').click()
        page.wait_for_load_state('networkidle', timeout=15000)

        all_rows = []
        week_cols = []

        print('[抓取各廠編]')
        for sid in PG_SUPPLIERS:
            brand = BRAND_MAP.get(sid, sid)
            print(f'  {sid} ({brand})...')
            wc, rows = fetch_supplier(page, sid)
            if not week_cols and wc:
                week_cols = wc
            for row in rows:
                if len(row) >= 6:
                    item_name = row[5] if len(row) > 5 else ''
                    all_rows.append({
                        'supplier_id': row[1] if len(row) > 1 else sid,
                        'brand': detect_brand(item_name, BRAND_MAP.get(sid, 'OTHER')),
                        'barcode': row[3] if len(row) > 3 else '',
                        'item_no': row[4] if len(row) > 4 else '',
                        'name': item_name,
                        'weeks': row[6:6+len(week_cols)]
                    })
            print(f'    {len(rows)} 筆')

        browser.close()

    # ── 寫 Excel ──────────────────────────────────
    print(f'[寫入 Excel] {out_path}')

    # 讀現有檔案或建新檔
    existing = os.path.expanduser('~/Downloads/康是美門市實銷資料(不包含網購 )-2026年04月.xlsx')
    if os.path.exists(existing):
        wb = openpyxl.load_workbook(existing)
    else:
        wb = openpyxl.Workbook()

    sheet_name = f'{today.year}.{today.month:02d}實銷'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    # 表頭
    hdr_fill = PatternFill('solid', fgColor='FF8C00')
    hdr_font = Font(bold=True, color='FFFFFF')
    headers = ['10碼廠編', '品牌', '條碼', '品號', '品名'] + week_cols
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    # 資料行
    for r, row in enumerate(all_rows, 2):
        ws.cell(r, 1, row['supplier_id'])
        ws.cell(r, 2, row['brand'])
        ws.cell(r, 3, row['barcode'])
        ws.cell(r, 4, row['item_no'])
        ws.cell(r, 5, row['name'])
        for w, val in enumerate(row['weeks'], 6):
            try:
                ws.cell(r, w, int(val.replace(',',''))) if val else None
            except:
                ws.cell(r, w, val)

    # 欄寬
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 45
    for col in range(6, 6+len(week_cols)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    wb.save(out_path)
    print(f'✅ 完成 — {len(all_rows)} 筆 SKU，{len(week_cols)} 週')
    print(f'   檔案: {out_path}')

if __name__ == '__main__':
    main()
