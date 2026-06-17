#!/usr/bin/env python3
"""
解析 iWMS 每日庫存單 xlsx（即期品）→ 更新 IWMS_EXPIRY
用法：python3 scripts/parse_iwms_upload.py <xlsx路徑>
      python3 scripts/parse_iwms_upload.py  （自動找 Downloads 最新檔）
"""
import sys, json, re, glob, os
from datetime import date, timedelta
from pathlib import Path
import openpyxl

DASHBOARD  = Path(__file__).parent.parent / 'dashboard.html'
BC_MAP_PATH = Path(__file__).parent.parent / 'data' / 'sku_barcode.json'

# 載入條碼對照表
_BC_MAP = {}
if BC_MAP_PATH.exists():
    _BC_MAP = json.load(open(BC_MAP_PATH, encoding='utf-8'))

def parse_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active

    # 找 header row（有「有效日期」的那行）
    hdr = None
    for i in range(1, 15):
        row = [str(sh.cell(i, j).value or '').strip() for j in range(1, 10)]
        if '有效日期' in row:
            hdr = i
            col_code = row.index('商品代碼') + 1 if '商品代碼' in row else 2
            col_name = row.index('商品名稱') + 1 if '商品名稱' in row else 3
            col_exp  = row.index('有效日期') + 1 if '有效日期' in row else 5
            col_qty  = row.index('庫存(SKU)') + 1 if '庫存(SKU)' in row else 6
            break
    if hdr is None:
        print('⚠ 找不到表頭，請確認是 iWMS 每日庫存單格式')
        return None

    # 讀取倉庫資訊（Row4 col3 = 物流中心名稱）
    wh_name = str(sh.cell(4, 3).value or '').strip() or '未知倉庫'

    today = date.today()
    d90  = today + timedelta(days=90)
    d180 = today + timedelta(days=180)
    expired, d90_list, d180_list, ok = [], [], [], []

    for i in range(hdr + 1, sh.max_row + 1):
        code = str(sh.cell(i, col_code).value or '').strip()
        name = str(sh.cell(i, col_name).value or '').strip()
        exp_raw = sh.cell(i, col_exp).value
        qty_raw = sh.cell(i, col_qty).value
        if not code or not name: continue
        if not code.startswith('8'): continue  # 非P&G品號（PB等）略過

        # 解析有效日期
        exp_str = ''
        if exp_raw:
            exp_str = str(exp_raw).strip()[:10]  # 取 YYYY-MM-DD
        if not exp_str or exp_str in ('None', ''):
            continue
        try:
            exp = date.fromisoformat(exp_str)
        except: continue

        qty = 0
        try: qty = float(qty_raw or 0)
        except: pass

        entry = {'code': code, 'bc': _BC_MAP.get(code, ''),
                 'name': name[:30], 'exp': exp_str,
                 'qty': qty, 'wh': wh_name}

        if   exp < today:   expired.append(entry)
        elif exp <= d90:    d90_list.append(entry)
        elif exp <= d180:   d180_list.append(entry)
        else:               ok.append(entry)

    print(f'  倉庫: {wh_name}')
    print(f'  已過期: {len(expired)} | 90天: {len(d90_list)} | 180天: {len(d180_list)} | 正常: {len(ok)}')
    return {
        'total': len(expired) + len(d90_list) + len(d180_list) + len(ok),
        'expired': len(expired),
        'd90': len(d90_list),
        'd180': len(d180_list),
        'ok': len(ok),
        'expired_top': sorted(expired,  key=lambda x: x['exp']),
        'd90_list':    sorted(d90_list,  key=lambda x: x['exp']),
        'd180_list':   sorted(d180_list, key=lambda x: x['exp']),
    }


def main():
    print('[iWMS 效期更新]')

    # 找檔案（支援多個路徑）
    if len(sys.argv) > 1:
        paths = sys.argv[1:]
    else:
        patterns = [
            os.path.expanduser('~/Downloads/每日庫存單*.xlsx'),
            os.path.expanduser('~/Downloads/iwms*.xlsx'),
        ]
        files = []
        for p in patterns:
            files.extend(glob.glob(p))
        if not files:
            print('✗ 找不到 iWMS 匯出檔，請手動指定路徑')
            sys.exit(1)
        paths = [max(files, key=os.path.getmtime)]

    warehouses = {}
    for path in paths:
        print(f'  來源: {os.path.basename(path)}')
        result = parse_file(path)
        if not result: continue
        wh_name = (result['expired_top'] or result['d90_list'] or result['d180_list'] or [{}])[0].get('wh', '未知倉庫')
        # 合計統計
        tot = result['total']; exp = result['expired']; d9 = result['d90']; d18 = result['d180']
        print(f'  → {wh_name}: {tot} 筆（過期:{exp} 90天:{d9} 180天:{d18}）')
        if wh_name in warehouses:
            # 同倉名（如高雄P&G全庫+即期倉）→ 合併
            ex = warehouses[wh_name]
            for k in ['expired_top','d90_list','d180_list']:
                ex[k] = sorted(ex[k] + result[k], key=lambda x: x['exp'])
            for k in ['total','expired','d90','d180','ok']:
                ex[k] = ex.get(k,0) + result.get(k,0)
        else:
            warehouses[wh_name] = result

    if not warehouses:
        print('✗ 沒有任何可用資料')
        sys.exit(1)

    today = str(date.today())
    data = {
        'updated': today,
        'warehouses': warehouses,
    }
    js = 'const IWMS_EXPIRY=' + json.dumps(data, ensure_ascii=False) + ';'

    html = DASHBOARD.read_text(encoding='utf-8')
    if 'const IWMS_EXPIRY=' in html:
        html = re.sub(r'const IWMS_EXPIRY=[\s\S]*?;', js, html)
    else:
        html = html.replace('</script>', f'\n{js}\n</script>', 1)
    DASHBOARD.write_text(html, encoding='utf-8')

    total_all = sum(w['total'] for w in warehouses.values())
    print(f'\n  ✅ IWMS_EXPIRY 寫入完成（{today}，{len(warehouses)} 倉，{total_all} 筆）')


if __name__ == '__main__':
    main()
